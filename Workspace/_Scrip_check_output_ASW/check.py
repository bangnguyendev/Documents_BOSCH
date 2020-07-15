import openpyxl
import sys

path_excel=str(sys.argv[1])
wb = openpyxl.load_workbook(path_excel, data_only=True)
sheet_Summary = wb['Summary']
print ("                C0 C1........................", sheet_Summary['A6'].value)
print ("                MCDC.........................", sheet_Summary['B6'].value)
print ("")
print ("    Reason : ", sheet_Summary['C6'].value)
print ("")
print ("===============")
try:
    sheet_Testcases = wb['Testcases']
except Exception as e:  
    sheet_Testcases = wb.worksheets[3]

'''Check input'''
TC_No = 1
Row_Tolerance = 18
Row_Type = 19
Row_Max = 20
Row_Min = 21
Row_Title = 22
Row_Name_Var = 23
Row_TC1 = 24
max_row_table = sheet_Testcases.max_row
max_column_table = sheet_Testcases.max_column
''''''
for i in range(Row_TC1,max_row_table):
    if str(sheet_Testcases.cell(i, TC_No).value) == 'None' :
        max_row_table = i
        break
     
'''Check TM name'''
if sheet_Testcases.cell(Row_Name_Var, TC_No).value is not None:
    pass
else:
    print (" ====> Error: [A23] Thieu TM_ name")


'''Read col input cont'''
print ("======== CHECK INPUT ")
for col in range(1,max_column_table+1):
    if str(sheet_Testcases.cell(Row_Title, col).value) == 'INPUTS':
        col_start_input = col
        break

for col in range(col_start_input + 1,max_column_table+1):
    if sheet_Testcases.cell(Row_Title, col).value is not None:
        col_end_input = col
        break

'''check row max min'''
for col in range(1,max_column_table+1):
    if str(sheet_Testcases.cell(Row_Title, col).value) == 'DESCRIPTIONS':
        col_end_row_max = col
        break
for col in range(col_start_input,col_end_row_max):
    if sheet_Testcases.cell(Row_Max, col).value is not None:
        type_value = str(sheet_Testcases.cell(Row_Type, col).value)
        if type_value == 'cont' or type_value == 'log' or type_value == 'enum':
            value_max = float(sheet_Testcases.cell(Row_Max, col).value)
            value_min = float(sheet_Testcases.cell(Row_Min, col).value)
            if value_max < value_min :
                print (" ====> Error: Check Row MaxMin ", str(sheet_Testcases.cell(Row_Name_Var, col).value))
        


for col in range(col_start_input,col_end_input):
    if str(sheet_Testcases.cell(Row_Type, col).value) == 'cont':
        if sheet_Testcases.cell(Row_Tolerance, col).value is not None:
            value_tol=float(sheet_Testcases.cell(Row_Tolerance, col).value)
        else:  
            value_tol = 'None'
            print (" ====> Error: Thieu Tolerance", str(sheet_Testcases.cell(Row_Name_Var, col).value))

        value_max=float(sheet_Testcases.cell(Row_Max, col).value)
        value_min=float(sheet_Testcases.cell(Row_Min, col).value)

        flag_not_out_max = 0
        flag_not_out_min = 0
        flag_not_mid_value = 0
        
        '''Check input out max'''
        for row in range(Row_TC1,max_row_table +1): 
            if sheet_Testcases.cell(row, col).value is not None:
                if float(sheet_Testcases.cell(row, col).value) > value_max :
                    flag_not_out_max = 0
                    break
                else :
                    flag_not_out_max = 1   
            else:
                break

        '''Check input out min'''
        for row in range(Row_TC1,max_row_table +1):   
            if sheet_Testcases.cell(row, col).value is not None:
                if float(sheet_Testcases.cell(row, col).value) < value_min :
                    flag_not_out_min = 0
                    break
                else :
                    flag_not_out_min = 1
            else:
                break
                
        '''Check input not mid value'''
        for row in range(Row_TC1,max_row_table +1):
            if sheet_Testcases.cell(row, col).value is not None:
                if float(sheet_Testcases.cell(row, col).value) < value_max and float(sheet_Testcases.cell(row, col).value) > value_min :
                    flag_not_mid_value = 0 
                    break
                else :
                    flag_not_mid_value = 1   
            else:
                break
        '''print resuit'''                
        if flag_not_out_max == 1:
            print (" ====> Error: None out max: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
        if flag_not_out_min == 1:
            print (" ====> Error: None out min: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
        if flag_not_mid_value == 1:
            print (" ====> Error: None mid value: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))

'''Check input log'''
for col in range(col_start_input,col_end_input):
    if str(sheet_Testcases.cell(Row_Type, col).value) == 'log':
        '''Check Tolerance'''
        try:
            value_tol=float(sheet_Testcases.cell(Row_Tolerance, col).value)
            if value_tol == 0 or value_tol == 1:
                pass
            else:
                print (" ====> Error: Tolerance failse", str(sheet_Testcases.cell(Row_Name_Var, col).value))
        except Exception as e:  
            value_tol = 'None'
            print (" ====> Error: Thieu Tolerance", str(sheet_Testcases.cell(Row_Name_Var, col).value))
        
        ''''''''''''''''''''''''''''''
        flag_true = 0   
        flag_false = 0     
        for row in range(Row_TC1,max_row_table +1):
            if sheet_Testcases.cell(row, col).value is not None:
                if str(sheet_Testcases.cell(row, col).value) == 'True':
                    flag_true = 1
                if str(sheet_Testcases.cell(row, col).value) == 'False':
                    flag_false = 1
            else:
                break
                
            if str(sheet_Testcases.cell(row, col).value) != 'True' and str(sheet_Testcases.cell(row, col).value) != 'False':
                print (" ====> Error: Sai dinh dang Bool ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
                
        if flag_true == 1 and flag_false == 1:
            pass
        else :
            print (" ====> Error: Thieu TRUE/FALSE: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
            
'''Check input enum'''
for col in range(col_start_input,col_end_input):
    if str(sheet_Testcases.cell(Row_Type, col).value) == 'enum':
        print ("WARNING Check enum min max mid for: ", str(sheet_Testcases.cell(Row_Name_Var, col).value))


'''Check variable as input '''
flag_yes_as_input = 0
for col in range(1,max_column_table+1):
    if str(sheet_Testcases.cell(Row_Title, col).value) == 'LOCAL VARIABLES AS INPUT':
        flag_yes_as_input = 1
        col_start_as_input = col
        break
if flag_yes_as_input == 1:
    print ("======== CHECK AS INPUT ")
    for col in range(col_start_as_input + 1,max_column_table+1):
        if sheet_Testcases.cell(Row_Title, col).value is not None:
            col_end_as_input = col
            break
    '''Check variable as input cont'''   
    for col in range(col_start_as_input,col_end_as_input):
        if str(sheet_Testcases.cell(Row_Type, col).value) == 'cont':
            if sheet_Testcases.cell(Row_Tolerance, col).value is not None:
                value_tol=float(sheet_Testcases.cell(Row_Tolerance, col).value)
            else: 
                value_tol = 'None'
                print (" ====> Error: Thieu Tolerance", str(sheet_Testcases.cell(Row_Name_Var, col).value))
            
            value_max=float(sheet_Testcases.cell(Row_Max, col).value)
            value_min=float(sheet_Testcases.cell(Row_Min, col).value)

            flag_max = 0
            flag_out_max = 0
            flag_min = 0
            flag_out_min = 0
            flag_mid_value = 0
            
            '''Check input max'''
            for row in range(Row_TC1,max_row_table +1): 
                if sheet_Testcases.cell(row, col).value is not None:
                    if float(sheet_Testcases.cell(row, col).value) == value_max :
                        flag_max = 1
      
                    if float(sheet_Testcases.cell(row, col).value) > value_max :
                        flag_out_max = 1
                else:
                    break

            '''Check input min'''
            for row in range(Row_TC1,max_row_table +1): 
                if sheet_Testcases.cell(row, col).value is not None:
                    if float(sheet_Testcases.cell(row, col).value) == value_min :
                        flag_min = 1

                    if float(sheet_Testcases.cell(row, col).value) < value_min :
                        flag_out_min = 1
                else:
                    break
     
            '''Check input not mid value'''
            for row in range(Row_TC1,max_row_table +1): 
                if sheet_Testcases.cell(row, col).value is not None:
                    if float(sheet_Testcases.cell(row, col).value) < value_max and float(sheet_Testcases.cell(row, col).value) > value_min :
                        flag_mid_value = 0 
                        break
                    else :
                        flag_mid_value = 1  
                else:
                    break

            '''Check '''
            flag_round_minmax = 0
            for row in range(Row_TC1,max_row_table+1): 
                if (sheet_Testcases.cell(row, col).value) is not None:
                    try:
                        float(sheet_Testcases.cell(row, col).value)
                        if float(sheet_Testcases.cell(row, col).value)%value_tol == 0 :
                            continue
                        else:
                            flag_round_minmax = 1
                        ## check lai voi max min    
                        if float(sheet_Testcases.cell(row, col).value) == value_max :
                            flag_round_minmax = 0
                        if float(sheet_Testcases.cell(row, col).value) == value_min :
                            flag_round_minmax = 0                                
                    except Exception as e:  
                        continue
                        '''print (" ====> WARNING: Div0", str(sheet_Testcases.cell(Row_Name_Var, col).value))'''
                else:
                    break

            '''print resuit'''          
            if flag_max == 0:
                print (" ====> Error: None value max: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))   
            if flag_min == 0:
                print (" ====> Error: None value min: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))     
            if flag_mid_value == 1:
                print (" ====> Error: None mid value: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))                
            if flag_out_max == 1:
                print (" ====> Error: Out range max: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
            if flag_out_min == 1:
                print (" ====> Error: Out range min: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
            if flag_round_minmax == 1:
                print (" ====> Warning: Check RoundMaxMin ",str(sheet_Testcases.cell(Row_Name_Var, col).value))

    '''Check variable as input log'''   
    for col in range(col_start_as_input,col_end_as_input):
        if str(sheet_Testcases.cell(Row_Type, col).value) == 'log':    
            flag_true = 0   
            flag_false = 0
            flag_format = 0
            '''check tolerance'''
            if sheet_Testcases.cell(Row_Tolerance, col).value is not None:
                value_tol=float(sheet_Testcases.cell(Row_Tolerance, col).value)
                if value_tol == 0 or value_tol == 1:
                    pass
                else:
                    print (" ====> Error: Tolerance failse", str(sheet_Testcases.cell(Row_Name_Var, col).value))
            else: 
                value_tol = 'None'
                print (" ====> Error: Thieu Tolerance", str(sheet_Testcases.cell(Row_Name_Var, col).value))
            
            ### check phai co true false        
            for row in range(Row_TC1,max_row_table+1):
                if sheet_Testcases.cell(row, col).value is not None:
                    if str(sheet_Testcases.cell(row, col).value) == 'True':
                        flag_true = 1
                    if str(sheet_Testcases.cell(row, col).value) == 'False':
                        flag_false = 1
                else:
                    break
                    
                if str(sheet_Testcases.cell(row, col).value) != 'True' and str(sheet_Testcases.cell(row, col).value) != 'False':
                    flag_format = 1
                                       
            if flag_true == 1 and flag_false == 1:
                pass
            else :
                print (" ====> Error: Thieu TRUE/FALSE: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
            
            if flag_format == 1:
                print (" ====> Error: Sai dinh dang Bool ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
           

'''Check IMPORTED PARAMETERS '''
flag_yes_imp_parm = 0
for col in range(1,max_column_table+1):
    if str(sheet_Testcases.cell(Row_Title, col).value) == 'IMPORTED PARAMETERS':
        print ("======== CHECK IMPORTED PARAMETERS ")
        flag_yes_imp_parm = 1
        col_start_imp_parm = col
        break
if flag_yes_imp_parm == 1:
    for col in range(col_start_imp_parm + 1,max_column_table+1):
        if sheet_Testcases.cell(Row_Title, col).value is not None:
            col_end_imp_parm = col
            break
    '''Check IMPORTED PARAMETERS cont'''        
    for col in range(col_start_imp_parm,col_end_imp_parm):
        if str(sheet_Testcases.cell(Row_Type, col).value) == 'cont':
            if str(sheet_Testcases.cell(Row_Min, col).value) == '-inf':
                pass
            elif str(sheet_Testcases.cell(Row_Min, col).value) == '-INF':
                pass
            else :               
                value_max=float(sheet_Testcases.cell(Row_Max, col).value)
                value_min=float(sheet_Testcases.cell(Row_Min, col).value)

                flag_max = 0
                flag_out_max = 0
                flag_min = 0
                flag_out_min = 0
                flag_mid_value = 0
                
                '''Check input max'''
                for row in range(Row_TC1,max_row_table+1): 
                    if sheet_Testcases.cell(row, col).value is not None:
                        if float(sheet_Testcases.cell(row, col).value) == value_max :
                            flag_max = 1
          
                        if float(sheet_Testcases.cell(row, col).value) > value_max :
                            flag_out_max = 1
                    else:
                        break
                   
                '''Check input min'''
                for row in range(Row_TC1,max_row_table+1): 
                    if sheet_Testcases.cell(row, col).value is not None:
                        if float(sheet_Testcases.cell(row, col).value) == value_min :
                            flag_min = 1

                        if float(sheet_Testcases.cell(row, col).value) < value_min :
                            flag_out_min = 1
                    else:
                        break
         
                '''Check input not mid value'''
                for row in range(Row_TC1,max_row_table+1): 
                    if sheet_Testcases.cell(row, col).value is not None:
                        if float(sheet_Testcases.cell(row, col).value) < value_max and float(sheet_Testcases.cell(row, col).value) > value_min :
                            flag_mid_value = 0 
                            break
                        else :
                            flag_mid_value = 1   
                    else:
                        break
                   
                '''print resuit'''          
                if flag_max == 0:
                    print (" ====> Error: None value max: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))   
                if flag_min == 0:
                    print (" ====> Error: None value min: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))     
                if flag_mid_value == 1:
                    print (" ====> Error: None mid value: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))                
                if flag_out_max == 1:
                    print (" ====> Error: Out range max: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
                if flag_out_min == 1:
                    print (" ====> Error: Out range min: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))

    '''Check IMPORTED PARAMETERS log'''   
    for col in range(col_start_imp_parm,col_end_imp_parm):
        if str(sheet_Testcases.cell(Row_Type, col).value) == 'log':    
            flag_true = 0   
            flag_false = 0
            flag_format = 0
            '''check tolerance'''
            if sheet_Testcases.cell(Row_Tolerance, col).value is not None:
                value_tol=float(sheet_Testcases.cell(Row_Tolerance, col).value)
                if value_tol == 0 or value_tol == 1:
                    pass
                else:
                    print (" ====> Error: Tolerance failse", str(sheet_Testcases.cell(Row_Name_Var, col).value))
            else: 
                value_tol = 'None'
                print (" ====> Error: Thieu Tolerance", str(sheet_Testcases.cell(Row_Name_Var, col).value))
            
            ### check phai co true false        
            for row in range(Row_TC1,max_row_table+1):
                if sheet_Testcases.cell(row, col).value is not None:
                    if str(sheet_Testcases.cell(row, col).value) == 'True':
                        flag_true = 1
                    if str(sheet_Testcases.cell(row, col).value) == 'False':
                        flag_false = 1
                else:
                    break
                    
                if str(sheet_Testcases.cell(row, col).value) != 'True' and str(sheet_Testcases.cell(row, col).value) != 'False':
                    flag_format = 1
                                       
            if flag_true == 1 and flag_false == 1:
                pass
            else :
                print (" ====> Error: Thieu TRUE/FALSE: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
            
            if flag_format == 1:
                print (" ====> Error: Sai dinh dang Bool ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
      
      
'''Check LOCAL VARIABLES '''            
flag_yes_lcal_var = 0
for col in range(1,max_column_table+1):
    if str(sheet_Testcases.cell(Row_Title, col).value) == 'LOCAL VARIABLES':
        print ("======== CHECK LOCAL VARIABLES ")
        flag_yes_lcal_var = 1
        col_start_lcal_var = col
        break
if flag_yes_lcal_var == 1:
    for col in range(col_start_lcal_var + 1,max_column_table+1):
        if sheet_Testcases.cell(Row_Title, col).value is not None:
            col_end_lcal_var = col
            break      

    for col in range(col_start_lcal_var,col_end_lcal_var):
        if str(sheet_Testcases.cell(Row_Type, col).value) == 'cont':
            if sheet_Testcases.cell(Row_Tolerance, col).value is not None:
                value_tol=float(sheet_Testcases.cell(Row_Tolerance, col).value)
            else:  
                value_tol = 'None'
                print (" ====> Error: Thieu Tolerance", str(sheet_Testcases.cell(Row_Name_Var, col).value))
    
        if str(sheet_Testcases.cell(Row_Type, col).value) == 'log':
            if sheet_Testcases.cell(Row_Tolerance, col).value is not None:
                value_tol=float(sheet_Testcases.cell(Row_Tolerance, col).value)
                if value_tol == 0 or value_tol == 1:
                    pass
                else:
                    print (" ====> Error: Tolerance failse", str(sheet_Testcases.cell(Row_Name_Var, col).value))
            else: 
                value_tol = 'None'
                print (" ====> Error: Thieu Tolerance", str(sheet_Testcases.cell(Row_Name_Var, col).value))

                

'''Check PARAMETERS '''            
flag_yes_parameters = 0
for col in range(1,max_column_table+1):
    if str(sheet_Testcases.cell(Row_Title, col).value) == 'PARAMETERS':
        print ("======== CHECK PARAMETERS ")
        flag_yes_parameters = 1
        col_start_parameters = col
        break
if flag_yes_parameters == 1:
    for col in range(col_start_parameters + 1,max_column_table+1):
        if sheet_Testcases.cell(Row_Title, col).value is not None:
            col_end_parameters = col
            break    
    '''Check PARAMETERS cont'''        
    for col in range(col_start_parameters,col_end_parameters):
        if str(sheet_Testcases.cell(Row_Type, col).value) == 'cont':                  
            if sheet_Testcases.cell(Row_Tolerance, col).value is not None:
                value_tol=float(sheet_Testcases.cell(Row_Tolerance, col).value)
            else:  
                value_tol = 'None'
                print (" ====> Error: Thieu Tolerance", str(sheet_Testcases.cell(Row_Name_Var, col).value))
            
            value_max=float(sheet_Testcases.cell(Row_Max, col).value)
            value_min=float(sheet_Testcases.cell(Row_Min, col).value)
            
            flag_out_max = 0
            flag_out_min = 0
            
            '''Check out range max'''
            for row in range(Row_TC1,max_row_table+1): 
                if (sheet_Testcases.cell(row, col).value) is not None:
                    try:
                        float(sheet_Testcases.cell(row, col).value)
                        if float(sheet_Testcases.cell(row, col).value) > value_max :
                            flag_out_max = 1
                            break
                    except Exception as e:  
                        continue
                        '''print (" ====> WARNING: Div0", str(sheet_Testcases.cell(Row_Name_Var, col).value))'''
                else:
                    break
            '''Check check out range min'''
            for row in range(Row_TC1,max_row_table+1): 
                if (sheet_Testcases.cell(row, col).value) is not None:
                    try:
                        float(sheet_Testcases.cell(row, col).value)
                        if float(sheet_Testcases.cell(row, col).value) < value_min :
                            flag_out_min = 1
                            break
                    except Exception as e:  
                        continue
                        '''print (" ====> WARNING: Div0", str(sheet_Testcases.cell(Row_Name_Var, col).value))'''
                else:
                    break
            '''Check '''
            flag_round_minmax = 0
            for row in range(Row_TC1,max_row_table+1): 
                if (sheet_Testcases.cell(row, col).value) is not None:
                    try:
                        float(sheet_Testcases.cell(row, col).value)
                        if float(sheet_Testcases.cell(row, col).value)%value_tol == 0 :
                            continue
                        else:
                            flag_round_minmax = 1
                        
                        ## check lai voi max min
                        if float(sheet_Testcases.cell(row, col).value) == value_max :
                            flag_round_minmax = 0
                        if float(sheet_Testcases.cell(row, col).value) == value_min :
                            flag_round_minmax = 0                        
                    except Exception as e:  
                        continue
                        '''print (" ====> WARNING: Div0", str(sheet_Testcases.cell(Row_Name_Var, col).value))'''
                else:
                    break
            '''print resuit'''          
            if flag_out_max == 1:
                print (" ====> Error: Out range max: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
            if flag_out_min == 1:
                print (" ====> Error: Out range min: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
            if flag_round_minmax == 1:
                print (" ====> Warning: Check RoundMaxMin ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
                        
'''Check OUTPUTS '''            
for col in range(1,max_column_table+1):
    if str(sheet_Testcases.cell(Row_Title, col).value) == 'OUTPUTS':
        print ("======== CHECK OUTPUTS ")
        col_start_outputs = col
        break

for col in range(col_start_outputs + 1,max_column_table+1):
    if sheet_Testcases.cell(Row_Title, col).value is not None:
        col_end_outputs = col
        break    

'''Check OUTPUTS cont'''        
for col in range(col_start_outputs,col_end_outputs):
####### check Tolerance
    if str(sheet_Testcases.cell(Row_Type, col).value) == 'cont':
        if sheet_Testcases.cell(Row_Tolerance, col).value is not None:
            value_tol=float(sheet_Testcases.cell(Row_Tolerance, col).value)
        else:  
            value_tol = 'None'
            print (" ====> Error: Thieu Tolerance", str(sheet_Testcases.cell(Row_Name_Var, col).value))

    if str(sheet_Testcases.cell(Row_Type, col).value) == 'log':
        if sheet_Testcases.cell(Row_Tolerance, col).value is not None:
            value_tol=float(sheet_Testcases.cell(Row_Tolerance, col).value)
            if value_tol == 0 or value_tol == 1:
                for row in range(Row_TC1,max_row_table+1):
                    if sheet_Testcases.cell(row, col).value is not None:
                        if str(sheet_Testcases.cell(row, col).value) != 'True' and str(sheet_Testcases.cell(row, col).value) != 'False':
                            print (" ====> Error: Sai dinh dang Bool ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
                    
            else:
                print (" ====> Error: Tolerance failse", str(sheet_Testcases.cell(Row_Name_Var, col).value))
        else: 
            value_tol = 'None'
            print (" ====> Error: Thieu Tolerance", str(sheet_Testcases.cell(Row_Name_Var, col).value))
####### check Max Min
    if str(sheet_Testcases.cell(Row_Type, col).value) == 'cont':                  

        value_max=float(sheet_Testcases.cell(Row_Max, col).value)
        value_min=float(sheet_Testcases.cell(Row_Min, col).value)

        flag_out_max = 0
        flag_out_min = 0
        
        '''Check output max'''
        for row in range(Row_TC1,max_row_table+1): 
            if sheet_Testcases.cell(row, col).value is not None:
                if float(sheet_Testcases.cell(row, col).value) > value_max :
                    flag_out_max = 1
                    break      
            else:
                break

        '''Check output min'''
        for row in range(Row_TC1,max_row_table+1): 
            if sheet_Testcases.cell(row, col).value is not None:
                if float(sheet_Testcases.cell(row, col).value) < value_min :
                    flag_out_min = 1
                    break
            else:
                break


        '''Check '''
        flag_round_minmax = 0
        for row in range(Row_TC1,max_row_table+1): 
            if (sheet_Testcases.cell(row, col).value) is not None:
                try:
                    float(sheet_Testcases.cell(row, col).value)
                    if float(sheet_Testcases.cell(row, col).value)%value_tol == 0 :
                        continue
                    else:
                        flag_round_minmax = 1
                    ## check lai voi max min    
                    if float(sheet_Testcases.cell(row, col).value) == value_max :
                        flag_round_minmax = 0
                    if float(sheet_Testcases.cell(row, col).value) == value_min :
                        flag_round_minmax = 0                                
                except Exception as e:  
                    continue
                    '''print (" ====> WARNING: Div0", str(sheet_Testcases.cell(Row_Name_Var, col).value))'''
            else:
                break
        '''print resuit'''          
        if flag_out_max == 1:
            print (" ====> Error: Out range max: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
        if flag_out_min == 1:
            print (" ====> Error: Out range min: ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
        if flag_round_minmax == 1:
            print (" ====> Warning: Check RoundMaxMin ",str(sheet_Testcases.cell(Row_Name_Var, col).value))
                        
wb.close()

 
