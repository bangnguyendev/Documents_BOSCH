import sys

import openpyxl

'''variable const'''
TC_No = 1
Row_Enum = 15
Row_Tolerance = 18
Row_Type = 19
Row_Max = 20
Row_Min = 21
Row_Title = 22
Row_Name_Var = 23
Row_TC1 = 24
color_Red = '\x1b[91m '
color_Green = '\x1b[92m '
color_Yellow = '\x1b[93m '
color_White = '\x1b[0m '
"""
17/7/2020
    Updated:    
        Check enum.
        Print color.
    Fixed:
        Khong can check tolerance cho imported parameter
        Cover them truong hop cho check_value_row_max_min()
        Fixed truong hop ngoai le cho file excel [long path, file khong ton tai.]
"""


def print_error(string=None, col_checking=None):
    """
    Print red color
    col_checking is type int.
    """
    if col_checking is not None:
        print(color_Red + string + color_White + str(sheet_TCs.cell(Row_Name_Var, col_checking).value))
    else:
        print(color_Red + string + color_White)
    return


def print_notice(string=None, col_checking=None):
    """
    Print green color
    """
    if col_checking is not None:
        print(color_Green + string + color_White + str(sheet_TCs.cell(Row_Name_Var, col_checking).value))
    else:
        print(color_Green + string + color_White)
    return


def print_warning(string=None, col_checking=None):
    """
    Print red yellow
    col_checking is type int.
    """
    if col_checking is not None:
        print(color_Yellow + string + color_White + str(sheet_TCs.cell(Row_Name_Var, col_checking).value))
    else:
        print(color_Yellow + string + color_White)
    return


try:
    ''' Neu co file '''
    path_excel = str(sys.argv[1])
    # path_excel = 'C:\\Users\\NguyenBang\\Desktop\\testassw\\MT_063_Psa_TCSPlus_MTCTargetSlip_DriverIntention
    # \\Delivery\\TestEnvironment\\Psa_TCSPlus_MTCTargetSlip_DriverIntention\\Documents\\TD_Psa_TCSPlus_MTCTargetSl
    # .xlsm'
    try:
        ''' Neu do dai link phu hop '''
        wb = openpyxl.load_workbook(path_excel, data_only=True)
        sheet_Summary = wb['Summary']
        print_notice("======= EXCEL TM_ File =======")
        print("                C0 C1........................", sheet_Summary['A6'].value)
        print("                MCDC.........................", sheet_Summary['B6'].value)
        print("")
        print("    Reason : ", sheet_Summary['C6'].value, "\n")
        print("========")

        # khai bao bien va chon sheet
        try:
            sheet_TCs = wb['Testcases']
        except:
            sheet_TCs = wb.worksheets[3]
    except:
        print_error(" ====> Error: Duong dan file dai. ")
        exit()
except:
    print_error(" ====> Error: Khong co file TD_*.xlsm. ")
    exit()

max_row_table = sheet_TCs.max_row
max_column_table = sheet_TCs.max_column

"""find number TCs"""
for row in range(Row_TC1, max_row_table + 1):
    if sheet_TCs.cell(row, TC_No).value is None:
        max_row_table = row
        break

"""find location range col input"""
col_start_input = 0
for col in range(1, max_column_table + 1):
    if str(sheet_TCs.cell(Row_Title, col).value) == 'INPUTS':
        col_start_input = col
        break

"""find location range col input"""
for col in range(col_start_input + 1, max_column_table + 1):
    if sheet_TCs.cell(Row_Title, col).value is not None:
        col_end_input = col
        break


def check_TM_name():
    """Check TM name"""
    if sheet_TCs.cell(Row_Name_Var, TC_No).value is not None:
        pass
    else:
        print_error("  ====> Error: [A23] Thieu TM_ name at sheet Testcases")
    return


def check_sum_TCs():
    """Check xem co bao nhieu TCs"""
    print_notice("===> Tong so TCs: " + str(max_row_table - Row_Name_Var))
    return


def check_value_row_max_min():
    """check row max min"""
    col_end_row_max = 0
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'DESCRIPTIONS':
            col_end_row_max = col
            break
    for col in range(col_start_input, col_end_row_max):
        flag_value_max = False
        flag_value_min = False
        if sheet_TCs.cell(Row_Max, col).value is not None:
            flag_value_max = True
        else:
            flag_value_max = False
        if sheet_TCs.cell(Row_Min, col).value is not None:
            flag_value_min = True
        else:
            flag_value_min = False

        type_value = str(sheet_TCs.cell(Row_Type, col).value)
        if type_value == 'cont' or type_value == 'log' or type_value == 'enum':
            if flag_value_max and flag_value_min:
                value_max = float(sheet_TCs.cell(Row_Max, col).value)
                value_min = float(sheet_TCs.cell(Row_Min, col).value)
                if value_max < value_min:
                    print_error(" ====> Error: Check Row Max < Min? ", col)
            elif flag_value_max and not flag_value_min:
                print_error(" ====> Error: Missing Row Min: ", col)
            elif not flag_value_max and flag_value_min:
                print_error(" ====> Error: Missing Row Max: ", col)
            else:
                print_error(" ====> Error: Missing Row Max-Min: ", col)
    """End check_value_row_max_min"""
    return


def check_input():
    """
    cont
        Check Tolerance YES/NO
        Check must be out max
        Check must be out min
        Check must be mid value
    log
        Check Tolerance 0:1
        Check must be has TRUE & FALSE
        Check fomat "TRUE/FALSE", not 1/0
    enum
        Print warning
        Check Tolerance
    """
    print_notice("======== CHECK INPUT ")
    for col in range(col_start_input, col_end_input):
        """check cont"""
        if str(sheet_TCs.cell(Row_Type, col).value) == 'cont':
            """Check Tolerance YES/NO"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
            else:
                value_tol = 'None'
                print_error(" ====> Error: Thieu Tolerance", col)

            """Check must be out max"""
            flag_ok_max = 0
            flag_ok_min = 0
            flag_not_out_max = 0
            flag_not_out_min = 0
            flag_not_mid_value = 0
            value_max = 0
            value_min = 0
            if sheet_TCs.cell(Row_Max, col).value is not None:
                value_max = float(sheet_TCs.cell(Row_Max, col).value)
                flag_ok_max = 1
            else:
                flag_ok_max = 0

            if sheet_TCs.cell(Row_Min, col).value is not None:
                value_min = float(sheet_TCs.cell(Row_Min, col).value)
                flag_ok_min = 1
            else:
                flag_ok_min = 0

            if flag_ok_min == 1 and flag_ok_max == 1:
                '''Check input out max'''
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if float(sheet_TCs.cell(row, col).value) > value_max:
                            flag_not_out_max = 0
                            break
                        else:
                            flag_not_out_max = 1
                    else:
                        break

                '''Check input out min'''
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if float(sheet_TCs.cell(row, col).value) < value_min:
                            flag_not_out_min = 0
                            break
                        else:
                            flag_not_out_min = 1
                    else:
                        break

                '''Check input not mid value'''
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if float(sheet_TCs.cell(row, col).value) < value_max and float(
                                sheet_TCs.cell(row, col).value) > value_min:
                            flag_not_mid_value = 0
                            break
                        else:
                            flag_not_mid_value = 1
                    else:
                        break
            else:
                print_error(" ====> Error: Missing row value Max/Min", col)

            '''print resuit'''
            if flag_not_out_max == 1:
                print_error(" ====> Error: None out max: ", col)
            if flag_not_out_min == 1:
                print_error(" ====> Error: None out min: ", col)
            if flag_not_mid_value == 1:
                print_error(" ====> Error: None mid value: ", col)

        """check log"""
        if str(sheet_TCs.cell(Row_Type, col).value) == 'log':
            """Check Tolerance YES/NO and must be 0 or 1"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                if isinstance(sheet_TCs.cell(Row_Tolerance, col).value, int):
                    value_tol = int(sheet_TCs.cell(Row_Tolerance, col).value)
                else:
                    print_error(" ====> Error:Row Tolerance must be 0 or 1", col)
            else:
                value_tol = 'None'
                print_error(" ====> Error: Thieu Tolerance", col)

            flag_true = 0
            "Check must be has TRUE & FALSE"
            flag_false = 0
            "Check must be has TRUE & FALSE"
            flag_format = 0
            "Check format TRUE/FALSE, not 1/0"
            for row in range(Row_TC1, max_row_table + 1):
                if sheet_TCs.cell(row, col).value is not None:
                    if str(sheet_TCs.cell(row, col).value) == 'True':
                        flag_true = 1
                    if str(sheet_TCs.cell(row, col).value) == 'False':
                        flag_false = 1
                else:
                    break

            for row in range(Row_TC1, max_row_table + 1):
                if sheet_TCs.cell(row, col).value is not None:
                    if str(sheet_TCs.cell(row, col).value) != 'True' and str(
                            sheet_TCs.cell(row, col).value) != 'False':
                        flag_format = 1
                        break

            if flag_true == 0 or flag_false == 0:
                print_error(" ====> Error: Thieu TRUE/FALSE: ", col)
            if flag_format == 1:
                print_error(" ====> Error: Wrong format Bool  ", col)

        '''Check input enum'''
        if str(sheet_TCs.cell(Row_Type, col).value) == 'enum':
            """Check Tolerance YES/NO and must be int"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                if value_tol == 0 or value_tol == 1:
                    pass
                else:
                    print_error(" ====> Error: Row Tolerance must be 'int' ", col)

                """Check max enum"""
                flag_ok_max = 0
                flag_ok_min = 0
                flag_enum_max = 0
                flag_enum_min = 0
                flag_enum_mid = 0
                flag_enum_format = 0
                value_max = 0
                value_min = 0
                number_of_enum = 0
                if sheet_TCs.cell(Row_Max, col).value is not None:
                    value_max = float(sheet_TCs.cell(Row_Max, col).value)
                    flag_ok_max = 1
                else:
                    flag_ok_max = 0

                if sheet_TCs.cell(Row_Min, col).value is not None:
                    value_min = float(sheet_TCs.cell(Row_Min, col).value)
                    flag_ok_min = 1
                else:
                    flag_ok_min = 0

                if flag_ok_min == 1 and flag_ok_max == 1:
                    '''Check enum max'''
                    " quet theo hang de in name enum "
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            for col_enum in range(TC_No, max_column_table + 1):
                                if isinstance(sheet_TCs.cell(row, col).value, str):
                                    name_enum = str(sheet_TCs.cell(row, col).value)
                                    if str(sheet_TCs.cell(Row_Enum, col_enum).value) == name_enum:
                                        number_of_enum = float(sheet_TCs.cell(Row_Enum + 1, col_enum).value)
                                        break
                                else:
                                    """Neu dinh dang sai thi bat co bao"""
                                    flag_enum_format = 1

                            if number_of_enum == value_max:
                                flag_enum_max = 1
                            if number_of_enum == value_min:
                                flag_enum_min = 1
                            if value_max > number_of_enum > value_min:
                                flag_enum_mid = 1
                        else:
                            break

                    if flag_enum_max == 0:
                        print_error(" ====> Error: Missing Enum max", col)

                    if flag_enum_min == 0:
                        print_error(" ====> Error: Missing Enum min", col)

                    if flag_enum_mid == 0:
                        print_error(" ====> Error: Missing Enum mid", col)

                    if flag_enum_format == 1:
                        print_error(" ====> Error: Wrong format Enum - Not is number ", col)

            else:
                value_tol = 'None'
                print_error(" ====> Error: Thieu Tolerance", col)

    """END check_input"""
    return


def check_as_input():
    """
    Check xem sheet nay co LOCAL VARIABLES AS INPUT hay khong?
    Neu co:
        cont
            Check Tolerance YES/NO
            Check must not be out max
            Check must not be out min
            Check must be mid value
        log
            Check Tolerance 0:1
            Check must be has TRUE & FALSE
            Check fomat "TRUE/FALSE", not 1/0
        enum
            Print warning
            Check Tolerance
    """
    flag_yes_as_input = 0
    "Flag check xem sheet co LOCAL VARIABLES AS INPUT hay khong?"
    col_start_as_input = 0
    col_end_as_input = 0
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'LOCAL VARIABLES AS INPUT':
            flag_yes_as_input = 1
            col_start_as_input = col
            break
    if flag_yes_as_input == 1:
        print_notice("======== CHECK AS INPUT ")
        for col in range(col_start_as_input + 1, max_column_table + 1):
            if sheet_TCs.cell(Row_Title, col).value is not None:
                col_end_as_input = col
                break
        "Check as input"
        for col in range(col_start_as_input, col_end_as_input):
            "Check cont"
            if str(sheet_TCs.cell(Row_Type, col).value) == 'cont':
                """Check Tolerance YES/NO"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

                """Check must be out max"""
                flag_ok_max = 0
                flag_ok_min = 0
                flag_max = 0
                flag_out_max = 0
                flag_min = 0
                flag_out_min = 0
                flag_mid_value = 0
                value_max = 0
                value_min = 0
                if sheet_TCs.cell(Row_Max, col).value is not None:
                    value_max = float(sheet_TCs.cell(Row_Max, col).value)
                    flag_ok_max = 1
                else:
                    flag_ok_max = 0

                if sheet_TCs.cell(Row_Min, col).value is not None:
                    value_min = float(sheet_TCs.cell(Row_Min, col).value)
                    flag_ok_min = 1
                else:
                    flag_ok_min = 0

                if flag_ok_min == 1 and flag_ok_max == 1:
                    '''Check input max'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            if float(sheet_TCs.cell(row, col).value) == value_max:
                                flag_max = 1

                            if float(sheet_TCs.cell(row, col).value) > value_max:
                                flag_out_max = 1
                        else:
                            break

                    '''Check input min'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            if float(sheet_TCs.cell(row, col).value) == value_min:
                                flag_min = 1

                            if float(sheet_TCs.cell(row, col).value) < value_min:
                                flag_out_min = 1
                        else:
                            break

                    '''Check input not mid value'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            if float(sheet_TCs.cell(row, col).value) < value_max and float(
                                    sheet_TCs.cell(row, col).value) > value_min:
                                flag_mid_value = 0
                                break
                            else:
                                flag_mid_value = 1
                        else:
                            break

                    '''print resuit'''
                    if flag_max == 0:
                        print_error(" ====> Error: None value max: ", col)
                    if flag_min == 0:
                        print_error(" ====> Error: None value min: ", col)
                    if flag_mid_value == 1:
                        print_error(" ====> Error: None mid value: ", col)
                    if flag_out_max == 1:
                        print_error(" ====> Error: Out range max: ", col)
                    if flag_out_min == 1:
                        print_error(" ====> Error: Out range min: ", col)
                else:
                    print_error(" ====> Error: Missing row value Max/Min", col)

            "Check log"
            if str(sheet_TCs.cell(Row_Type, col).value) == 'log':
                """Check Tolerance YES/NO and must be 0 or 1"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                    if value_tol == 0 or value_tol == 1:
                        pass
                    else:
                        print_error(" ====> Error: Row Tolerance must be 0 or 1", col)
                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

                flag_true = 0
                "Check must be has TRUE & FALSE"
                flag_false = 0
                "Check must be has TRUE & FALSE"
                flag_format = 0
                "Check format TRUE/FALSE, not 1/0"
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if str(sheet_TCs.cell(row, col).value) == 'True':
                            flag_true = 1
                        if str(sheet_TCs.cell(row, col).value) == 'False':
                            flag_false = 1
                    else:
                        break

                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if str(sheet_TCs.cell(row, col).value) != 'True' and str(
                                sheet_TCs.cell(row, col).value) != 'False':
                            flag_format = 1
                            break

                if flag_true == 0 or flag_false == 0:
                    print_error(" ====> Error: Thieu TRUE/FALSE: ", col)
                if flag_format == 1:
                    print_error(" ====> Error: Wrong format Bool  ", col)
            '''Check input enum'''
            if str(sheet_TCs.cell(Row_Type, col).value) == 'enum':
                """Check Tolerance YES/NO and must be int"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    if isinstance(sheet_TCs.cell(Row_Tolerance, col).value, int):
                        value_tol = int(sheet_TCs.cell(Row_Tolerance, col).value)
                    else:
                        print_error(" ====> Error: Row Tolerance must be 'int' ", col)

                    """Check max enum"""
                    flag_ok_max = 0
                    flag_ok_min = 0
                    flag_enum_max = 0
                    flag_enum_min = 0
                    flag_enum_mid = 0
                    flag_enum_format = 0
                    value_max = 0
                    value_min = 0
                    number_of_enum = 0
                    if sheet_TCs.cell(Row_Max, col).value is not None:
                        value_max = float(sheet_TCs.cell(Row_Max, col).value)
                        flag_ok_max = 1
                    else:
                        flag_ok_max = 0

                    if sheet_TCs.cell(Row_Min, col).value is not None:
                        value_min = float(sheet_TCs.cell(Row_Min, col).value)
                        flag_ok_min = 1
                    else:
                        flag_ok_min = 0
                    '''Check enum max min mid'''
                    if flag_ok_min == 1 and flag_ok_max == 1:
                        for row in range(Row_TC1, max_row_table + 1):
                            if sheet_TCs.cell(row, col).value is not None:
                                for col_enum in range(TC_No, max_column_table + 1):
                                    if isinstance(sheet_TCs.cell(row, col).value, str):
                                        name_enum = str(sheet_TCs.cell(row, col).value)
                                        if str(sheet_TCs.cell(Row_Enum, col_enum).value) == name_enum:
                                            number_of_enum = float(sheet_TCs.cell(Row_Enum + 1, col_enum).value)
                                            break
                                    else:
                                        """Neu dinh dang sai thi bat co bao"""
                                        flag_enum_format = 1

                                if number_of_enum == value_max:
                                    flag_enum_max = 1
                                if number_of_enum == value_min:
                                    flag_enum_min = 1
                                if value_max > number_of_enum > value_min:
                                    flag_enum_mid = 1
                            else:
                                break

                        if flag_enum_max == 0:
                            print_error(" ====> Error: Missing Enum max", col)
                        if flag_enum_min == 0:
                            print_error(" ====> Error: Missing Enum min", col)
                        if flag_enum_mid == 0:
                            print_error(" ====> Error: Missing Enum mid", col)
                        if flag_enum_format == 1:
                            print_error(" ====> Error: Wrong format Enum - Not is number ", col)
                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

    """END check_as_input"""
    return


def check_imported_parameters():
    """
    Check xem co IMPORTED PARAMETERS hay khong?
    Neu co:
        cont
            Check Tolerance
            inf -inf number
            Check must not be out max
            Check must not be out min
            Check must be mid value
            Check when value is const
        log
            Check Tolerance 0:1
            Check must be has TRUE & FALSE
            Check fomat "TRUE/FALSE", not 1/0
        enum
            Print warning
            Check Tolerance
    :return:
    """
    flag_yes_imp_parm = 0
    "Flag check xem co IMPORTED PARAMETERS hay khong"
    col_start_imp_parm = 0
    col_end_imp_parm = 0
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'IMPORTED PARAMETERS':
            flag_yes_imp_parm = 1
            col_start_imp_parm = col
            break
    if flag_yes_imp_parm == 1:
        print_notice("======== CHECK IMPORTED PARAMETERS ")
        for col in range(col_start_imp_parm + 1, max_column_table + 1):
            if sheet_TCs.cell(Row_Title, col).value is not None:
                col_end_imp_parm = col
                break
        '''Check IMPORTED PARAMETERS '''
        for col in range(col_start_imp_parm, col_end_imp_parm):
            '''Check IMPORTED PARAMETERS cont'''
            if str(sheet_TCs.cell(Row_Type, col).value) == 'cont':
                if str(sheet_TCs.cell(Row_Min, col).value) == '-inf':
                    pass
                elif str(sheet_TCs.cell(Row_Min, col).value) == '-INF':
                    pass
                elif str(sheet_TCs.cell(Row_Min, col).value) == '-Inf':
                    pass
                else:
                    # """Check Tolerance YES/NO"""
                    # if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    # value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                    # else:
                    # value_tol = 'None'
                    # print_error(" ====> Error: Thieu Tolerance", col)

                    """Check must be out max"""
                    flag_ok_max = 0
                    flag_ok_min = 0
                    flag_max = 0
                    flag_out_max = 0
                    flag_min = 0
                    flag_out_min = 0
                    flag_mid_value = 0
                    value_max = 0
                    value_min = 0
                    if sheet_TCs.cell(Row_Max, col).value is not None:
                        value_max = float(sheet_TCs.cell(Row_Max, col).value)
                        flag_ok_max = 1
                    else:
                        flag_ok_max = 0

                    if sheet_TCs.cell(Row_Min, col).value is not None:
                        value_min = float(sheet_TCs.cell(Row_Min, col).value)
                        flag_ok_min = 1
                    else:
                        flag_ok_min = 0

                    if flag_ok_max == 1 and flag_ok_min == 1:
                        '''Check input max'''
                        for row in range(Row_TC1, max_row_table + 1):
                            if sheet_TCs.cell(row, col).value is not None:
                                if float(sheet_TCs.cell(row, col).value) == value_max:
                                    flag_max = 1

                                if float(sheet_TCs.cell(row, col).value) > value_max:
                                    flag_out_max = 1
                            else:
                                break

                        '''Check input min'''
                        for row in range(Row_TC1, max_row_table + 1):
                            if sheet_TCs.cell(row, col).value is not None:
                                if float(sheet_TCs.cell(row, col).value) == value_min:
                                    flag_min = 1

                                if float(sheet_TCs.cell(row, col).value) < value_min:
                                    flag_out_min = 1
                            else:
                                break

                        '''Check input not mid value'''
                        for row in range(Row_TC1, max_row_table + 1):
                            if sheet_TCs.cell(row, col).value is not None:
                                if float(sheet_TCs.cell(row, col).value) < value_max and float(
                                        sheet_TCs.cell(row, col).value) > value_min:
                                    flag_mid_value = 0
                                    break
                                else:
                                    flag_mid_value = 1
                            else:
                                break

                        '''Check xem co phai const hay khong? '''
                        if value_max == value_min:
                            flag_mid_value = 0

                        '''print resuit'''
                        if flag_max == 0:
                            print_error(" ====> Error: None value max: ", col)
                        if flag_min == 0:
                            print_error(" ====> Error: None value min: ", col)
                        if flag_mid_value == 1:
                            print_error(" ====> Error: None mid value: ", col)
                        if flag_out_max == 1:
                            print_error(" ====> Error: Out range max: ", col)
                        if flag_out_min == 1:
                            print_error(" ====> Error: Out range min: ", col)
                    else:
                        print_error(" ====> Error: Missing row value Max/Min", col)

            "Check log"
            if str(sheet_TCs.cell(Row_Type, col).value) == 'log':
                """Check Tolerance YES/NO and must be 0 or 1"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                    if value_tol == 0 or value_tol == 1:
                        pass
                    else:
                        print_error(" ====> Error: Row Tolerance must be 0 or 1", col)
                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

                flag_true = 0
                "Check must be has TRUE & FALSE"
                flag_false = 0
                "Check must be has TRUE & FALSE"
                flag_format = 0
                "Check format TRUE/FALSE, not 1/0"
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if str(sheet_TCs.cell(row, col).value) == 'True':
                            flag_true = 1
                        if str(sheet_TCs.cell(row, col).value) == 'False':
                            flag_false = 1
                    else:
                        break

                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if str(sheet_TCs.cell(row, col).value) != 'True' and str(
                                sheet_TCs.cell(row, col).value) != 'False':
                            flag_format = 1
                            break

                if flag_true == 0 or flag_false == 0:
                    print_error(" ====> Error: Thieu TRUE/FALSE: ", col)
                if flag_format == 1:
                    print_error(" ====> Error: Wrong format Bool  ", col)
            '''Check input enum'''
            if str(sheet_TCs.cell(Row_Type, col).value) == 'enum':
                """Check Tolerance YES/NO and must be int"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    if isinstance(sheet_TCs.cell(Row_Tolerance, col).value, int):
                        value_tol = int(sheet_TCs.cell(Row_Tolerance, col).value)
                    else:
                        print_error(" ====> Error: Row Tolerance must be 'int' ", col)

                    """Check max enum"""
                    flag_ok_max = 0
                    flag_ok_min = 0
                    flag_enum_max = 0
                    flag_enum_min = 0
                    flag_enum_mid = 0
                    flag_enum_format = 0
                    value_max = 0
                    value_min = 0
                    number_of_enum = 0
                    if sheet_TCs.cell(Row_Max, col).value is not None:
                        value_max = float(sheet_TCs.cell(Row_Max, col).value)
                        flag_ok_max = 1
                    else:
                        flag_ok_max = 0

                    if sheet_TCs.cell(Row_Min, col).value is not None:
                        value_min = float(sheet_TCs.cell(Row_Min, col).value)
                        flag_ok_min = 1
                    else:
                        flag_ok_min = 0
                    '''Check enum max min mid'''
                    if flag_ok_min == 1 and flag_ok_max == 1:
                        for row in range(Row_TC1, max_row_table + 1):
                            if sheet_TCs.cell(row, col).value is not None:
                                for col_enum in range(TC_No, max_column_table + 1):
                                    if isinstance(sheet_TCs.cell(row, col).value, str):
                                        name_enum = str(sheet_TCs.cell(row, col).value)
                                        if str(sheet_TCs.cell(Row_Enum, col_enum).value) == name_enum:
                                            number_of_enum = float(sheet_TCs.cell(Row_Enum + 1, col_enum).value)
                                            break
                                    else:
                                        """Neu dinh dang sai thi bat co bao"""
                                        flag_enum_format = 1

                                if number_of_enum == value_max:
                                    flag_enum_max = 1
                                if number_of_enum == value_min:
                                    flag_enum_min = 1
                                if value_max > number_of_enum > value_min:
                                    flag_enum_mid = 1
                            else:
                                break

                        if flag_enum_max == 0:
                            print_error(" ====> Error: Missing Enum max", col)
                        if flag_enum_min == 0:
                            print_error(" ====> Error: Missing Enum min", col)
                        if flag_enum_mid == 0:
                            print_error(" ====> Error: Missing Enum mid", col)
                        if flag_enum_format == 1:
                            print_error(" ====> Error: Wrong format Enum - Not is number ", col)

                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

    """END check IMPORTED PARAMETERS"""
    return


def check_local_variable():
    """
    Check xem co LOCAL VARIABLES hay khong?
    Neu co:
        cont
            Check Tolenrance
            Check must not be out max
            Check must not be out min
            Check must be mid value
    :return:
    """
    flag_yes_lcal_var = 0
    '''Flag check xem co LOCAL VARIABLES hay khong?'''
    col_start_lcal_var = 0
    col_end_lcal_var = 0
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'LOCAL VARIABLES':
            flag_yes_lcal_var = 1
            col_start_lcal_var = col
            break
    if flag_yes_lcal_var == 1:
        print_notice("======== CHECK LOCAL VARIABLES ")
        for col in range(col_start_lcal_var + 1, max_column_table + 1):
            if sheet_TCs.cell(Row_Title, col).value is not None:
                col_end_lcal_var = col
                break
        '''Check LOCAL VARIABLES'''
        for col in range(col_start_lcal_var, col_end_lcal_var):
            '''check cont'''
            if str(sheet_TCs.cell(Row_Type, col).value) == 'cont':
                """Check Tolerance YES/NO"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

                """Check must be out max"""
                flag_ok_max = 0
                flag_ok_min = 0
                flag_max = 0
                flag_out_max = 0
                flag_min = 0
                flag_out_min = 0
                flag_mid_value = 0
                flag_check_round = 0
                flag_div0 = 0
                value_max = 0
                value_min = 0
                if sheet_TCs.cell(Row_Max, col).value is not None:
                    value_max = float(sheet_TCs.cell(Row_Max, col).value)
                    flag_ok_max = 1
                else:
                    flag_ok_max = 0

                if sheet_TCs.cell(Row_Min, col).value is not None:
                    value_min = float(sheet_TCs.cell(Row_Min, col).value)
                    flag_ok_min = 1
                else:
                    flag_ok_min = 0

                if flag_ok_min == 1 and flag_ok_max == 1:
                    '''Check input max'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            try:
                                if float(sheet_TCs.cell(row, col).value) == value_max:
                                    flag_max = 1

                                if float(sheet_TCs.cell(row, col).value) > value_max:
                                    flag_out_max = 1
                            except ValueError:
                                flag_div0 = 1
                        else:
                            break

                    '''Check input min'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            try:
                                if float(sheet_TCs.cell(row, col).value) == value_min:
                                    flag_min = 1

                                if float(sheet_TCs.cell(row, col).value) < value_min:
                                    flag_out_min = 1
                            except ValueError:
                                flag_div0 = 1
                        else:
                            break

                    '''Check input not mid value'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            try:
                                if float(sheet_TCs.cell(row, col).value) < value_max and float(
                                        sheet_TCs.cell(row, col).value) > value_min:
                                    flag_mid_value = 0  # co mid value
                                    break
                                else:
                                    flag_mid_value = 1  # khong co mid value
                            except ValueError:
                                flag_div0 = 1
                        else:
                            break

                    '''Check xem khong co mid value co phu hop hay khong?'''
                    if flag_mid_value == 1:  # khong co mid value
                        for row in range(Row_TC1, max_row_table + 1):
                            if sheet_TCs.cell(row, col).value is not None:
                                try:
                                    if float(sheet_TCs.cell(row, col).value) == value_max:
                                        flag_check_round = 0
                                    elif float(sheet_TCs.cell(row, col).value) == value_min:
                                        flag_check_round = 0
                                    else:
                                        flag_check_round = 1
                                        break

                                except ValueError:
                                    flag_div0 = 1
                            else:
                                break

                    '''print resuit'''
                    if flag_max == 0:
                        print_error(" ====> Error: None value max: ", col)
                    if flag_min == 0:
                        print_error(" ====> Error: None value min: ", col)
                    if flag_check_round == 1:
                        print_error(" ====> Error: Check round min max : ", col)
                    if flag_out_max == 1:
                        print_error(" ====> Error: Out range max: ", col)
                    if flag_out_min == 1:
                        print_error(" ====> Error: Out range min: ", col)
                    if flag_div0 == 1:
                        print_error(" ====> Error: Must not be 'DIV/0' : ", col)

                else:
                    print_error(" ====> Error: Missing row value Max/Min", col)
    '''END check local param'''
    return


def check_parameters():
    """
    Check xem co PARAMETERS hay khong?
    Neu co:
        cont
            Check Tolenrance
            Check must not be out max
            Check must not be out min
            Check must be mid value
        log
            Check Tolerance 0:1
            Check must be has TRUE & FALSE
            Check fomat "TRUE/FALSE", not 1/0
    :return:
    """
    flag_yes_parameters = 0
    '''Flag check xem co PARAMETERS hay khong? '''
    col_start_parameters = 0
    col_end_parameters = 0
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'PARAMETERS':
            print_notice("======== CHECK PARAMETERS ")
            flag_yes_parameters = 1
            col_start_parameters = col
            break
    if flag_yes_parameters == 1:
        for col in range(col_start_parameters + 1, max_column_table + 1):
            if sheet_TCs.cell(Row_Title, col).value is not None:
                col_end_parameters = col
                break
        '''Check PARAMETERS'''
        for col in range(col_start_parameters, col_end_parameters):
            '''Check cont'''
            if str(sheet_TCs.cell(Row_Type, col).value) == 'cont':
                """Check Tolerance YES/NO"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

                """Check must be out max"""
                flag_ok_max = 0
                flag_ok_min = 0
                flag_out_max = 0
                flag_out_min = 0
                flag_div0 = 0
                value_max = 0
                value_min = 0
                if sheet_TCs.cell(Row_Max, col).value is not None:
                    value_max = float(sheet_TCs.cell(Row_Max, col).value)
                    flag_ok_max = 1
                else:
                    flag_ok_max = 0

                if sheet_TCs.cell(Row_Min, col).value is not None:
                    value_min = float(sheet_TCs.cell(Row_Min, col).value)
                    flag_ok_min = 1
                else:
                    flag_ok_min = 0

                if flag_ok_min == 1 and flag_ok_max == 1:
                    '''Check input max'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            try:
                                if float(sheet_TCs.cell(row, col).value) > value_max:
                                    flag_out_max = 1
                            except ValueError:
                                flag_div0 = 1
                        else:
                            break

                    '''Check input min'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            try:
                                if float(sheet_TCs.cell(row, col).value) < value_min:
                                    flag_out_min = 1
                            except ValueError:
                                flag_div0 = 1
                        else:
                            break

                    '''print resuit'''
                    if flag_out_max == 1:
                        print_error(" ====> Error: Out range max: ", col)
                    if flag_out_min == 1:
                        print_error(" ====> Error: Out range min: ", col)
                    if flag_div0 == 1:
                        print_warning(" ====> WARNING: DIV/0: ", col)

                else:
                    print_error(" ====> Error: Missing row value Max/Min", col)
            '''Check log'''
            if str(sheet_TCs.cell(Row_Type, col).value) == 'log':
                """Check Tolerance YES/NO and must be 0 or 1"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                    if value_tol == 0 or value_tol == 1:
                        pass
                    else:
                        print_error(" ====> Error: Row Tolerance must be 0 or so nguyen ", col)
                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

                flag_true = 0
                "Check must be has TRUE & FALSE"
                flag_false = 0
                "Check must be has TRUE & FALSE"
                flag_format = 0
                "Check format TRUE/FALSE, not 1/0"
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if str(sheet_TCs.cell(row, col).value) == 'True':
                            flag_true = 1
                        if str(sheet_TCs.cell(row, col).value) == 'False':
                            flag_false = 1
                    else:
                        break

                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if str(sheet_TCs.cell(row, col).value) != 'True' and str(
                                sheet_TCs.cell(row, col).value) != 'False':
                            flag_format = 1
                            break

                if flag_true == 0 and flag_false == 0:
                    print_error(" ====> Error: Thieu TRUE/FALSE: ", col)
                if flag_format == 1:
                    print_error(" ====> Error: Wrong format Bool  ", col)
    '''END check PARAMETERS'''
    return


def check_output():
    """
    cont
        Check Tolerance
        Check must not be out max
        Check must not be out min
        Check must be mid value
        Check when value is const
    log
        Check Tolerance 0:1
        Check must be has TRUE & FALSE
        Check fomat "TRUE/FALSE", not 1/0
    enum
        Check Tolerance
        Check enum not type number
    :return:
    """

    print_notice("======== CHECK OUTPUTS ")
    col_start_outputs = 0
    col_end_outputs = 0
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'OUTPUTS':
            col_start_outputs = col
            break

    for col in range(col_start_outputs + 1, max_column_table + 1):
        if sheet_TCs.cell(Row_Title, col).value is not None:
            col_end_outputs = col
            break
    '''Check OUTPUT'''
    for col in range(col_start_outputs, col_end_outputs):
        '''Check cont'''
        if str(sheet_TCs.cell(Row_Type, col).value) == 'cont':
            """Check Tolerance YES/NO"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
            else:
                value_tol = 'None'
                print_error(" ====> Error: Thieu Tolerance", col)

            """Check must be out max"""
            flag_ok_max = 0
            flag_ok_min = 0
            flag_out_max = 0
            flag_out_min = 0
            flag_div0 = 0
            value_max = 0
            value_min = 0
            if sheet_TCs.cell(Row_Max, col).value is not None:
                value_max = float(sheet_TCs.cell(Row_Max, col).value)
                flag_ok_max = 1
            else:
                flag_ok_max = 0

            if sheet_TCs.cell(Row_Min, col).value is not None:
                value_min = float(sheet_TCs.cell(Row_Min, col).value)
                flag_ok_min = 1
            else:
                flag_ok_min = 0

            if flag_ok_min == 1 and flag_ok_max == 1:
                '''Check input max'''
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        try:
                            if float(sheet_TCs.cell(row, col).value) > value_max:
                                flag_out_max = 1
                        except ValueError:
                            flag_div0 = 1
                    else:
                        break

                '''Check input min'''
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        try:
                            if float(sheet_TCs.cell(row, col).value) < value_min:
                                flag_out_min = 1
                        except ValueError:
                            flag_div0 = 1
                    else:
                        break

                '''print resuit'''
                if flag_out_max == 1:
                    print_error(" ====> Error: Out range max: ", col)
                if flag_out_min == 1:
                    print_error(" ====> Error: Out range min: ", col)
                if flag_div0 == 1:
                    print_warning(" ====> WARNING: DIV/0: ", col)
            else:
                print_error(" ====> Error: Missing row value Max/Min", col)
        '''Check log'''
        if str(sheet_TCs.cell(Row_Type, col).value) == 'log':
            """Check Tolerance YES/NO and must be 0 or 1"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                if value_tol == 0 or value_tol == 1:
                    pass
                else:
                    print_error(" ====> Error: Row Tolerance must be 0 or 1", col)
            else:
                value_tol = 'None'
                print_error(" ====> Error: Thieu Tolerance", col)

            flag_true = 0
            "Check must be has TRUE or FALSE"
            flag_false = 0
            "Check must be has TRUE or FALSE"
            flag_format = 0
            "Check format TRUE/FALSE, not 1/0"
            for row in range(Row_TC1, max_row_table + 1):
                if sheet_TCs.cell(row, col).value is not None:
                    if str(sheet_TCs.cell(row, col).value) == 'True':
                        flag_true = 1
                    if str(sheet_TCs.cell(row, col).value) == 'False':
                        flag_false = 1
                else:
                    break

            for row in range(Row_TC1, max_row_table + 1):
                if sheet_TCs.cell(row, col).value is not None:
                    if str(sheet_TCs.cell(row, col).value) != 'True' and str(
                            sheet_TCs.cell(row, col).value) != 'False':
                        flag_format = 1
                        break

            if flag_true == 0 and flag_false == 0:
                print_error(" ====> Error: Thieu TRUE/FALSE: ", col)
            if flag_format == 1:
                print_error(" ====> Error: Wrong format Bool  ", col)
        '''Check input enum'''
        if str(sheet_TCs.cell(Row_Type, col).value) == 'enum':
            """Check Tolerance YES/NO and must be int"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                if isinstance(sheet_TCs.cell(Row_Tolerance, col).value, int):
                    value_tol = int(sheet_TCs.cell(Row_Tolerance, col).value)
                else:
                    print_error(" ====> Error: Row Tolerance must be 'int' ", col)

                flag_ok_max = 0
                flag_ok_min = 0
                flag_enum_format = 0

                if sheet_TCs.cell(Row_Max, col).value is not None:
                    flag_ok_max = 1
                else:
                    flag_ok_max = 0

                if sheet_TCs.cell(Row_Min, col).value is not None:
                    flag_ok_min = 1
                else:
                    flag_ok_min = 0
                '''Check enum format'''
                if flag_ok_min == 1 and flag_ok_max == 1:
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            for col_enum in range(TC_No, max_column_table + 1):
                                if not isinstance(sheet_TCs.cell(row, col).value, str):
                                    """Neu dinh dang sai thi bat co bao"""
                                    flag_enum_format = 1
                        else:
                            break

                    if flag_enum_format == 1:
                        print_error(" ====> Error: Wrong format Enum - Not is number ", col)
            else:
                value_tol = 'None'
                print_error(" ====> Error: Thieu Tolerance", col)
    '''End check output'''
    return


check_TM_name()
check_sum_TCs()
check_value_row_max_min()
check_input()
check_as_input()
check_imported_parameters()
check_local_variable()
check_parameters()
check_output()
