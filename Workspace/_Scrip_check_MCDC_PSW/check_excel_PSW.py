import sys
import openpyxl
import win32com.client
from win32com.client import Dispatch


# '''variable const'''
name_user = 'EXTERNAL Nguyen Duy Bang (Ban Vien, RBVH/EPS45)'

color_Red = '\x1b[91m '
color_Green = '\x1b[92m '
color_Yellow = '\x1b[93m '
color_Orange = '\x1b[38;5;208m '
color_White = '\x1b[0m '
E_OK = True
E_NOT_OK = False


xl = win32com.client.Dispatch('Excel.Application')
wb = xl.Workbooks.Open('C:\\_BangNguyen\\documents_bosch\\Workspace\\_Scrip_check_MCDC_PSW\\NET_Adaptive_DLC_NIS_CodeCoverage_or_Fail_Reason.xls')
readData = wb.WorkSheets('Unit_Test_Info')
allData = readData.UsedRange 

# Get number of rows used on active sheet
getNumRows = allData.Rows.Count
# print ('Number of rows used in sheet : ',getNumRows)

#Get number of columns used on active sheet
getNumCols = allData.Columns.Count
# print ('Number of columns used in sheet : ',getNumCols)

def get_value(row=0, col=0):
    return allData.Cells(row,col).value

for row in range(1, getNumRows +1):
    for col in range(1, getNumCols +1):
        if get_value(row,col) == "Tester":
            row_tester = row
            col_tester = col
            print("Tester: ",get_value(row,col +1))
        elif get_value(row,col) == "Date":
            row_date = row
            col_date = col
            print("Date: ", get_value(row,col +1))
        elif get_value(row,col) == "Item_Name":
            row_Item_Name = row
            col_Item_Name = col
            print("Item_Name: ", get_value(row,col +1))
        elif get_value(row,col) == "C0":
            row_C0 = row
            col_C0 = col
            print(get_value(row,col +1))
            if float(get_value(row,col +1)) < 100:
                print("aaaaaaaaaaaaaaaaaaaa11111")
        elif get_value(row,col) == "C1":
            row_C1 = row
            col_C1 = col
            print(get_value(row,col +1))
            if float(get_value(row,col +1)) < 100:
                print("aaaaaaaaaaaaaaaaaaaa2222")
        elif get_value(row,col) == "MCDC":
            row_MCDC = row
            col_MCDC = col
            print(get_value(row,col +1))
            if float(get_value(row,col +1)) < 100:
                print("aaaaaaaaaaaaaaaaaaaa333")

# print(allData.Cells(row_tester,col_tester).value)


def print_error(string=None, col_checking=None):
    """
    Print red color
    col_checking is type int.
    """
    if col_checking is not None:
        print(color_Red + string + color_White + str(sheet_TCs.cell(Row_Name_Var, col_checking).value))
    else:
        print(color_Red + string + color_White)
    return


def print_notice(string=None, col_checking=None):
    """
    Print green color
    col_checking is type int.
    """
    if col_checking is not None:
        print(color_Green + string + color_White + str(sheet_TCs.cell(Row_Name_Var, col_checking).value))
    else:
        print(color_Green + string + color_White)
    return


def print_warning(string=None, col_checking=None):
    """
    Print yellow color
    col_checking is type int.
    """
    if col_checking is not None:
        print(color_Yellow + string + color_White + str(sheet_TCs.cell(Row_Name_Var, col_checking).value))
    else:
        print(color_Yellow + string + color_White)
    return

wb.Close()
# 1