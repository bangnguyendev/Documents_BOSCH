from openpyxl import load_workbook
import xlsxwriter

wb = load_workbook(filename='Summary_JOEM_COEM_20200601.xlsm', data_only = 'True')
# wb = load_workbook(filename='okok.xlsx')
# ws = wb['Sheet1']
ws = wb['Merged_COEM']
# ws = wb['Merged_JOEM']
print(ws)

var_max_column = ws.max_column
var_max_row = ws.max_row
# tim vi tri table
row_table = 0
for i in range(1, var_max_column):
    for j in range(1, var_max_row):
        if ws.cell(row=j, column=i).value == "TaskID":
            row_table = j
            print("row_table:", row_table)
# tim vi tri ItemName
ItemName = 0
for i in range(1, var_max_column):
    if ws.cell(row=row_table, column=i).value == "ItemName":
        ItemName = i
        print("ItemName:", ItemName)
# tim vi tri Status
Status = 0
for i in range(1, var_max_column):
    if ws.cell(row=row_table, column=i).value == "Status":
        Status = i
        print("Status:", Status)
# tim vi tri Start
Start = 0
for i in range(1, var_max_column):
    if ws.cell(row=row_table, column=i).value == "Start":
        Start = i
        print("Start:", Start)
# tim vi tri End
End = 0
for i in range(1, var_max_column):
    if ws.cell(row=row_table, column=i).value == "End":
        End = i
        print("End:", End)
# tim vi tri Status_Result
Status_Result = 0
for i in range(1, var_max_column):
    if ws.cell(row=row_table, column=i).value == "Status Result":
        Status_Result = i
        print("Status_Result:", Status_Result)
# tim vi tri C0
C0 = 0
for i in range(1, var_max_column):
    if ws.cell(row=row_table, column=i).value == "C0":
        C0 = i
        print("C0:", C0)
# tim vi tri C1
C1 = 0
for i in range(1, var_max_column):
    if ws.cell(row=row_table, column=i).value == "C1":
        C1 = i
        print("C1:", C1)
# tim vi tri MCDC
MCDC = 0
for i in range(1, var_max_column):
    if ws.cell(row=row_table, column=i).value == "MCDC":
        MCDC = i
        print("MCDC:", MCDC)

# tim vi tri hang chua ten "nguyen duy bang"
for i in range(row_table, var_max_row):
    for j in range(1, var_max_column):
        if ws.cell(row=i, column=j).value == "bang.nguyen-duy":
            print("Row", i, ":", ws.cell(row=i, column=j).value, ws.cell(row=i, column=ItemName).value)


wb.close()
