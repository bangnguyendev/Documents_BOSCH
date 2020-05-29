from openpyxl import load_workbook, Workbook

wb = load_workbook(filename = 'Summary_JOEM_COEM_20200601.xlsm')

ws = wb.active

var_max_column = ws.max_column
var_max_row = ws.max_row
# tim vi tri table
for i in range(1,var_max_column):
    for j in range(1,var_max_row):
        if ws.cell(row=j, column=i).value == "TaskID" :
            row_table = j
            print("row_table:",row_table)
# tim vi tri ItemName
for i in range(1,var_max_column):
    if ws.cell(row=row_table, column=i).value == "ItemName" :
        ItemName = i
        print("ItemName:",ItemName)
# tim vi tri Status
for i in range(1,var_max_column):
    if ws.cell(row=row_table, column=i).value == "Status" :
        Status = i
        print("Status:",Status)
# tim vi tri Start
for i in range(1,var_max_column):
    if ws.cell(row=row_table, column=i).value == "Start" :
        Start = i
        print("Start:",Start)
# tim vi tri End
for i in range(1,var_max_column):
    if ws.cell(row=row_table, column=i).value == "End" :
        End = i
        print("End:",End)
# tim vi tri Status_Result
for i in range(1,var_max_column):
    if ws.cell(row=row_table, column=i).value == "Status Result" :
        Status_Result = i
        print("Status_Result:",Status_Result)
# tim vi tri C0
for i in range(1,var_max_column):
    if ws.cell(row=row_table, column=i).value == "C0" :
        C0 = i
        print("C0:",C0)
# tim vi tri C1
for i in range(1,var_max_column):
    if ws.cell(row=row_table, column=i).value == "C1" :
        C1 = i
        print("C1:",C1)
# tim vi tri MCDC
for i in range(1, var_max_column):
    if ws.cell(row=row_table, column=i).value == "MCDC":
        MCDC = i
        print("MCDC:",MCDC)

# tim vi tri hang chua ten "nguyen duy bang"
for i in range(row_table,var_max_row):
    for j in range(1,var_max_column):
        if ws.cell(row=i, column=j).value == "bang.nguyen-duy" :
            print("Row", i, ":", ws.cell(row=i, column=j).value, ws.cell(row=i, column=ItemName).value)




